"""
Utility functions for Extractor Jurisprudencias.
Provides logging, config management, network fetching with retries, HTML parsing,
and high-fidelity TSJ official print format PDF generation.
"""

import os
import json
import logging
import re
import time
import unicodedata
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import requests
import urllib3
from bs4 import BeautifulSoup
import pandas as pd
from colorama import init, Fore, Style

# Disable SSL verification warnings for TSJ Venezuela server
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT

# Initialize colorama for Windows terminal
init(autoreset=True)

# Map of room identifiers to full names and TSJ codes
SALAS_MAP = {
    "constitucional": {
        "nombre": "Sala Constitucional",
        "code": "scon",
        "alias": ["constitucional", "scon", "sala constitucional"]
    },
    "politico_administrativa": {
        "nombre": "Sala Político-Administrativa",
        "code": "spa",
        "alias": ["politico_administrativa", "spa", "politico", "politica"]
    },
    "casacion_civil": {
        "nombre": "Sala de Casación Civil",
        "code": "scc",
        "alias": ["casacion_civil", "scc", "civil"]
    },
    "casacion_penal": {
        "nombre": "Sala de Casación Penal",
        "code": "scp",
        "alias": ["casacion_penal", "scp", "penal"]
    },
    "casacion_social": {
        "nombre": "Sala de Casación Social",
        "code": "scs",
        "alias": ["casacion_social", "scs", "social"]
    },
    "electoral": {
        "nombre": "Sala Electoral",
        "code": "se",
        "alias": ["electoral", "se", "sala electoral"]
    },
    "plena": {
        "nombre": "Sala Plena",
        "code": "plena",
        "alias": ["plena", "sala plena"]
    }
}

# Choices for interactive TSJ Chamber and Month selection
SALAS_CHOICES = {
    "0": {"key": "todas", "nombre": "Todas las Salas / Juzgados", "code": "todas"},
    "1": {"key": "plena", "nombre": "Sala Plena", "code": "plena"},
    "2": {"key": "constitucional", "nombre": "Sala Constitucional", "code": "scon"},
    "3": {"key": "politico_administrativa", "nombre": "Sala Político-Administrativa", "code": "spa"},
    "4": {"key": "electoral", "nombre": "Sala Electoral", "code": "se"},
    "5": {"key": "casacion_civil", "nombre": "Sala de Casación Civil", "code": "scc"},
    "6": {"key": "casacion_penal", "nombre": "Sala de Casación Penal", "code": "scp"},
    "7": {"key": "casacion_social", "nombre": "Sala de Casación Social", "code": "scs"}
}

MESES_CHOICES = {
    "0": {"key": "todo_el_ano", "nombre": "Todo el Año en curso (Enero - Diciembre)", "code": "ano_completo"},
    "1": {"key": "enero", "nombre": "Enero", "code": "enero"},
    "2": {"key": "febrero", "nombre": "Febrero", "code": "febrero"},
    "3": {"key": "marzo", "nombre": "Marzo", "code": "marzo"},
    "4": {"key": "abril", "nombre": "Abril", "code": "abril"},
    "5": {"key": "mayo", "nombre": "Mayo", "code": "mayo"},
    "6": {"key": "junio", "nombre": "Junio", "code": "junio"},
    "7": {"key": "julio", "nombre": "Julio", "code": "julio"},
    "8": {"key": "agosto", "nombre": "Agosto", "code": "agosto"},
    "9": {"key": "septiembre", "nombre": "Septiembre", "code": "septiembre"},
    "10": {"key": "octubre", "nombre": "Octubre", "code": "octubre"},
    "11": {"key": "noviembre", "nombre": "Noviembre", "code": "noviembre"},
    "12": {"key": "diciembre", "nombre": "Diciembre", "code": "diciembre"}
}


def prompt_select_sala() -> Dict[str, str]:
    """Displays DC3 style interactive menu for selecting TSJ Chamber."""
    print(f"\n{Fore.BLUE}{'='*78}")
    print(f"{Fore.YELLOW}  [ TSJ CYBER FORENSICS LAB ] {Fore.CYAN}SELECCIÓN DE SALA O JUZGADO TSJ")
    print(f"{Fore.BLUE}{'='*78}{Style.RESET_ALL}")
    print(f"  {Fore.GREEN}[0]{Style.RESET_ALL} TODAS LAS SALAS (Plena, Constitucional, Político-Admin, Electoral, Civil, Penal, Social)")
    print(f"  {Fore.YELLOW}[1]{Style.RESET_ALL} Sala Plena")
    print(f"  {Fore.YELLOW}[2]{Style.RESET_ALL} Sala Constitucional")
    print(f"  {Fore.YELLOW}[3]{Style.RESET_ALL} Sala Político-Administrativa")
    print(f"  {Fore.YELLOW}[4]{Style.RESET_ALL} Sala Electoral")
    print(f"  {Fore.YELLOW}[5]{Style.RESET_ALL} Sala de Casación Civil")
    print(f"  {Fore.YELLOW}[6]{Style.RESET_ALL} Sala de Casación Penal")
    print(f"  {Fore.YELLOW}[7]{Style.RESET_ALL} Sala de Casación Social\n")
    
    opc = input(f"{Fore.YELLOW}Seleccione el número de Sala [0-7] (Por defecto: 0): {Style.RESET_ALL}").strip()
    return SALAS_CHOICES.get(opc, SALAS_CHOICES["0"])


def prompt_select_mes() -> Dict[str, str]:
    """Displays DC3 style interactive menu for selecting Month / Year period."""
    print(f"\n{Fore.BLUE}{'='*78}")
    print(f"{Fore.YELLOW}  [ TSJ CYBER FORENSICS LAB ] {Fore.CYAN}SELECCIÓN DE MES O PERÍODO A ESCANEAR")
    print(f"{Fore.BLUE}{'='*78}{Style.RESET_ALL}")
    print(f"  {Fore.GREEN}[0]{Style.RESET_ALL} TODO EL AÑO EN CURSO (Enero a Diciembre)")
    print(f"  {Fore.YELLOW}[1]{Style.RESET_ALL} Enero      {Fore.YELLOW}[2]{Style.RESET_ALL} Febrero    {Fore.YELLOW}[3]{Style.RESET_ALL} Marzo      {Fore.YELLOW}[4]{Style.RESET_ALL} Abril")
    print(f"  {Fore.YELLOW}[5]{Style.RESET_ALL} Mayo       {Fore.YELLOW}[6]{Style.RESET_ALL} Junio      {Fore.YELLOW}[7]{Style.RESET_ALL} Julio      {Fore.YELLOW}[8]{Style.RESET_ALL} Agosto")
    print(f"  {Fore.YELLOW}[9]{Style.RESET_ALL} Septiembre {Fore.YELLOW}[10]{Style.RESET_ALL} Octubre   {Fore.YELLOW}[11]{Style.RESET_ALL} Noviembre  {Fore.YELLOW}[12]{Style.RESET_ALL} Diciembre\n")
    
    opc = input(f"{Fore.YELLOW}Seleccione el número de Mes [0-12] (Por defecto: 0): {Style.RESET_ALL}").strip()
    return MESES_CHOICES.get(opc, MESES_CHOICES["0"])


def get_canonical_filenames(sala_key: str = "todas") -> Tuple[str, str, str]:
    """
    Returns canonical SQLite DB path, Excel path, and clean name for a given Sala.
    Standardized format: sala_<nombre_limpio>.db and sala_<nombre_limpio>.xlsx
    """
    sala_clean_map = {
        "todas": "tsj_todas_las_salas",
        "constitucional": "sala_constitucional",
        "politico_administrativa": "sala_politico_administrativa",
        "casacion_civil": "sala_casacion_civil",
        "casacion_penal": "sala_casacion_penal",
        "casacion_social": "sala_casacion_social",
        "electoral": "sala_electoral",
        "plena": "sala_plena"
    }
    clean_name = sala_clean_map.get(sala_key, "tsj_todas_las_salas")
    db_path = os.path.join("data/Databases_SQLite", f"{clean_name}.db")
    excel_path = os.path.join("data/Excel_Buscador", f"{clean_name}.xlsx")
    return db_path, excel_path, clean_name


# Keywords specifically for digital evidence, computer crimes, and technology
TECH_KEYWORDS = [
    "prueba digital", "evidencia digital", "delito informático", "delitos informáticos",
    "ciberdelito", "delito cibernético", "peritaje informático", "tecnología",
    "telecomunicaciones", "datos informáticos", "correo electrónico", "redes sociales",
    "mensajería digital", "intercepción de comunicaciones", "firma electrónica",
    "documento digital", "seguridad informática", "cadena de custodia digital",
    "acceso indebido", "fraude electrónico", "sabotaje informático", "whatsapp", "software",
    "evidencia electrónica", "pericia informática", "delitos de tecnología"
]


def setup_logger(name: str = "ExtractorJurisprudencia", log_file: Optional[str] = "extractor.log") -> logging.Logger:
    """Configures and returns a logger with colored console output and file logging."""
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)
    
    if logger.handlers:
        return logger

    # Console Handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    class ColoredFormatter(logging.Formatter):
        FORMATS = {
            logging.DEBUG: Fore.CYAN + "[DEBUG] %(message)s" + Style.RESET_ALL,
            logging.INFO: Fore.GREEN + "[INFO] %(message)s" + Style.RESET_ALL,
            logging.WARNING: Fore.YELLOW + "[WARN] %(message)s" + Style.RESET_ALL,
            logging.ERROR: Fore.RED + "[ERROR] %(message)s" + Style.RESET_ALL,
            logging.CRITICAL: Fore.RED + Style.BRIGHT + "[CRITICAL] %(message)s" + Style.RESET_ALL,
        }

        def format(self, record):
            log_fmt = self.FORMATS.get(record.levelno, "%(message)s")
            formatter = logging.Formatter(log_fmt)
            return formatter.format(record)

    console_handler.setFormatter(ColoredFormatter())
    logger.addHandler(console_handler)

    # File Handler
    if log_file:
        os.makedirs(os.path.dirname(log_file) if os.path.dirname(log_file) else ".", exist_ok=True)
        file_handler = logging.FileHandler(log_file, encoding="utf-8")
        file_handler.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")
        file_handler.setFormatter(file_formatter)
        logger.addHandler(file_handler)

    return logger


def load_config(config_path: str = "config.json") -> Dict[str, Any]:
    """Loads configuration from JSON file with default fallbacks."""
    default_config = {
        "sala": "Sala Constitucional",
        "sala_id": "constitucional",
        "ano": 2024,
        "palabra_clave": "",
        "limite": 10,
        "formato_salida": "pdf",
        "carpeta_salida": "data/PDFs",
        "descargar_pdf_directo": True,
        "modo_tecnologia": False,
        "escanear_todas_las_salas": False,
        "carpeta_tecnologia": "data/Jurisprudencia_Tecnologica",
        "timeout": 15,
        "max_retries": 3
    }
    
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                user_config = json.load(f)
                default_config.update(user_config)
        except Exception as e:
            print(f"{Fore.YELLOW}[WARN] Error loading {config_path}: {e}. Using defaults.{Style.RESET_ALL}")
            
    return default_config


def save_config(config_data: Dict[str, Any], config_path: str = "config.json") -> None:
    """Saves dictionary configuration to JSON file."""
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config_data, f, indent=2, ensure_ascii=False)


def sanitize_search_name(query_or_formula: str, prefix: str = "Busqueda") -> str:
    """
    Sanitizes search query/formula into a clean, filesystem-safe string
    used for naming both SQLite databases (.db) and Excel workbooks (.xlsx).
    Example: 'Delitos Informáticos (2024)' -> 'Busqueda_delitos_informaticos_2024'
    """
    if not query_or_formula or not str(query_or_formula).strip():
        return f"{prefix}_General"
    
    # Normalize unicode (strip accents)
    nfkd_form = unicodedata.normalize('NFKD', str(query_or_formula))
    text = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    
    # Replace non-alphanumeric characters with underscores
    clean = re.sub(r'[^a-zA-Z0-9]+', '_', text).strip('_')
    
    if not clean:
        return f"{prefix}_General"
        
    if prefix and not clean.lower().startswith(prefix.lower()):
        return f"{prefix}_{clean}"
        
    return clean



def clean_text(text: Optional[str]) -> str:
    """Cleans up raw HTML text strings by stripping excess whitespace."""
    if not text:
        return ""
    text = re.sub(r'\r\n|\r|\n', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def parse_jurisprudencia_html(html_content: str, url: str = "") -> Dict[str, Any]:
    """
    Parses HTML text of a decision to extract structured TSJ metadata.
    Fields extracted: numero_sentencia, fecha, ponente, expediente, materia, tema, asunto, resumen, texto_completo
    """
    soup = BeautifulSoup(html_content, "lxml")
    
    # Text content extraction
    raw_text = soup.get_text(separator="\n")
    cleaned_lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
    full_text = "\n\n".join(cleaned_lines)
    
    # Metadata extraction using Regex patterns
    num_match = re.search(r'(?:sentencia|Nº|N°|Nro\.?|Número)\s*[:\.]?\s*(\d+[-–]\d+|\d+)', full_text, re.IGNORECASE)
    numero_sentencia = num_match.group(1) if num_match else "S/N"
    
    exp_match = re.search(r'(?:expediente|exp\.?|Nº?\s*de?\s*exp\.?)\s*[:\.]?\s*([A-Z0-9\-\.]{4,})', full_text, re.IGNORECASE)
    expediente = exp_match.group(1) if exp_match else "S/E"
    
    ponente_match = re.search(r'(?:magistrad[ao]\s+ponente|ponente)\s*[:\.]?\s*((?:Dr[a]?\.\s*)?[A-ZÁÉÍÓÚÑa-záéíóúñ\'\s]+)', full_text, re.IGNORECASE)
    ponente = clean_text(ponente_match.group(1)) if ponente_match else "No especificado"
    
    materia_match = re.search(r'(?:materia)\s*[:\.]?\s*([A-ZÁÉÍÓÚÑa-záéíóúñ\s]+?)(?:\n|asunto|tema|$)', full_text, re.IGNORECASE)
    materia = clean_text(materia_match.group(1)) if materia_match else "Derecho Constitucional / Penal"

    tema_match = re.search(r'(?:tema)\s*[:\.]?\s*([A-ZÁÉÍÓÚÑa-záéíóúñ\s]+?)(?:\n|asunto|materia|$)', full_text, re.IGNORECASE)
    tema = clean_text(tema_match.group(1)) if tema_match else "Garantías Procesales y Evidencia Digital"

    fecha_match = re.search(r'(\d{1,2}\s+de\s+[a-z]+?\s+de\s+\d{4})', full_text, re.IGNORECASE)
    fecha = fecha_match.group(1) if fecha_match else datetime.now().strftime("%d de %B de %Y")
    
    resumen = full_text[:400] + "..." if len(full_text) > 400 else full_text
    
    return {
        "numero_sentencia": numero_sentencia,
        "expediente": expediente,
        "fecha": fecha,
        "ponente": ponente,
        "materia": materia,
        "tema": tema,
        "url": url,
        "resumen": resumen,
        "texto_completo": full_text
    }


def export_to_json(data: List[Dict[str, Any]], filepath: str) -> None:
    """Exports list of decision dictionaries to JSON format."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def export_to_csv(data: List[Dict[str, Any]], filepath: str) -> None:
    """Exports list of decision dictionaries to CSV format."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df = pd.DataFrame(data)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")


def export_to_excel(data: List[Dict[str, Any]], filepath: str) -> None:
    """Exports list of decision dictionaries to Excel format."""
    export_tsj_to_excel_profesional(data, filepath)


def export_tsj_to_excel_profesional(data: List[Dict[str, Any]], filepath: str, title: str = "Búsqueda e Indización de Jurisprudencia TSJ") -> None:
    """
    Exports TSJ decisions to a professionally styled Excel workbook:
    - Navy Blue (#1A365D) institutional headers with white bold text
    - Direct clickable HTTP/HTTPS hyperlinks for each TSJ decision URL
    - Structured 9 columns (Sala, Nº Sentencia, Nº Expediente, Fecha, Tema, Materia, Asunto, Extracto, Link Directo)
    - Fine borders, custom alignments, and auto-adjusted column widths for rapid reading
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jurisprudencias Indexadas"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Sala TSJ",
        "Nº Sentencia",
        "Nº Expediente",
        "Fecha",
        "Tema",
        "Materia",
        "Asunto / Síntesis",
        "Extracto (Detalle Pop-up)",
        "Link Directo TSJ"
    ]
    
    ws.append(headers)

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D2D6DC'),
        right=Side(style='thin', color='D2D6DC'),
        top=Side(style='thin', color='D2D6DC'),
        bottom=Side(style='thin', color='D2D6DC')
    )

    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for r_idx, dec in enumerate(data, 2):
        sala = str(dec.get("sala", "Sala Constitucional")).replace("Sala de ", "").replace("Sala ", "")
        sent = str(dec.get("numero_sentencia", dec.get("sentencia", "S/N")))
        exp = str(dec.get("expediente", "S/E"))
        fecha = str(dec.get("fecha", "N/A"))
        tema = str(dec.get("tema", "Sentencia"))
        materia = str(dec.get("materia", "Derecho Procesal"))
        asunto = str(dec.get("asunto", ""))
        extracto = str(dec.get("extracto", dec.get("asunto", "")))
        url = str(dec.get("link_directo", dec.get("url", dec.get("link", "https://historico.tsj.gob.ve"))))

        ws.cell(row=r_idx, column=1, value=f"Sala {sala}").alignment = center_align
        ws.cell(row=r_idx, column=2, value=sent).alignment = center_align
        ws.cell(row=r_idx, column=3, value=exp).alignment = center_align
        ws.cell(row=r_idx, column=4, value=fecha).alignment = center_align
        ws.cell(row=r_idx, column=5, value=tema).alignment = left_align
        ws.cell(row=r_idx, column=6, value=materia).alignment = left_align
        ws.cell(row=r_idx, column=7, value=asunto).alignment = left_align
        ws.cell(row=r_idx, column=8, value=extracto).alignment = left_align

        # Clickable Hyperlink Cell
        url_cell = ws.cell(row=r_idx, column=9, value=url)
        url_cell.hyperlink = url
        url_cell.font = Font(name="Arial", size=10, color="0000FF", underline="single")
        url_cell.alignment = left_align

        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).border = thin_border

    column_widths = [24, 14, 18, 20, 26, 24, 55, 65, 70]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filepath)




def generate_decision_pdf(decision: Dict[str, Any], filepath: str) -> None:
    """
    Generates a high-fidelity PDF reproducing the official TSJ Venezuela print format.
    Matches the exact layout of TSJ web decisions:
    - Official Escudo de Venezuela Seal header
    - Official TSJ heading text in Times-Roman Bold
    - Magistrado Ponente line
    - TSJ metadata card box (Expediente, Sentencia, Tema, Materia, Asunto, Link)
    - Full judgment text with centered legal subheadings
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        rightMargin=45,
        leftMargin=45,
        topMargin=35,
        bottomMargin=35
    )
    
    styles = getSampleStyleSheet()
    
    # TSJ Official Header Styles (Times-Roman)
    tsj_rep_style = ParagraphStyle(
        'TSJRep',
        fontName='Times-Bold',
        fontSize=13,
        leading=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )
    
    tsj_sub_rep_style = ParagraphStyle(
        'TSJSubRep',
        fontName='Times-Bold',
        fontSize=10,
        leading=13,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )
    
    tsj_tsj_style = ParagraphStyle(
        'TSJTitle',
        fontName='Times-Bold',
        fontSize=12,
        leading=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )
    
    tsj_sala_style = ParagraphStyle(
        'TSJSala',
        fontName='Times-Bold',
        fontSize=12,
        leading=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )
    
    tsj_ponente_style = ParagraphStyle(
        'TSJPonente',
        fontName='Times-Bold',
        fontSize=10.5,
        leading=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000")
    )
    
    box_title_style = ParagraphStyle(
        'BoxTitle',
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#1A202C")
    )
    
    box_body_style = ParagraphStyle(
        'BoxBody',
        fontName='Helvetica',
        fontSize=9.5,
        leading=13.5,
        textColor=colors.HexColor("#2D3748")
    )
    
    body_style = ParagraphStyle(
        'TSJBodyText',
        parent=styles['Normal'],
        fontName='Times-Roman',
        fontSize=10.5,
        leading=15,
        alignment=TA_JUSTIFY,
        textColor=colors.HexColor("#000000")
    )

    section_heading_style = ParagraphStyle(
        'TSJSectionHeading',
        fontName='Times-Bold',
        fontSize=11,
        leading=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#000000"),
        spaceBefore=10,
        spaceAfter=6
    )
    
    story = []
    
    # 1. Official Escudo Emblem
    escudo_path = "assets/escudo_venezuela.png"
    if not os.path.exists(escudo_path):
        from src.create_assets import generate_escudo_image
        generate_escudo_image(escudo_path)

    if os.path.exists(escudo_path):
        img = Image(escudo_path, width=70, height=75)
        img.hAlign = 'CENTER'
        story.append(img)
        story.append(Spacer(1, 6))

    # 2. Official TSJ Header Text
    story.append(Paragraph("REPÚBLICA BOLIVARIANA DE VENEZUELA", tsj_rep_style))
    story.append(Paragraph("EN SU NOMBRE", tsj_sub_rep_style))
    story.append(Paragraph("EL TRIBUNAL SUPREMO DE JUSTICIA", tsj_tsj_style))
    
    sala_name = decision.get('sala', 'SALA CONSTITUCIONAL').upper()
    if not sala_name.startswith("SALA"):
        sala_name = f"SALA {sala_name}"
    story.append(Paragraph(sala_name, tsj_sala_style))
    story.append(Spacer(1, 12))
    
    # 3. Magistrado Ponente Line
    ponente_text = decision.get('ponente', 'ELSA JANETH GÓMEZ MORENO')
    if not ponente_text.lower().startswith("magistrad"):
        ponente_text = f"Magistrado Ponente Doctor {ponente_text}"
    story.append(Paragraph(ponente_text, tsj_ponente_style))
    story.append(Spacer(1, 14))
    
    # 4. TSJ Official Decision Metadata Box (Matching TSJ Web Print Box)
    box_content = [
        [Paragraph(f"<b>Nº de Expediente: {decision.get('expediente', 'N/A')}</b> &nbsp;&nbsp;&nbsp;&nbsp; <b>Nº de Sentencia: {decision.get('numero_sentencia', 'N/A')}</b>", box_title_style)],
        [Paragraph(f"<b>Tema:</b> {decision.get('tema', decision.get('asunto', 'Motivación y Evidencia Digital'))}", box_body_style)],
        [Paragraph(f"<b>Materia:</b> {decision.get('materia', 'Derecho Procesal y Delitos Tecnológicos')}", box_body_style)],
        [Paragraph(f"<b>Asunto:</b> {decision.get('asunto', 'Análisis de validez de la prueba digital y licitud del peritaje informático.')}", box_body_style)],
        [Paragraph(f"<b>Ver Extracto:</b> <font color='#2B6CB0'><u>{decision.get('url', 'http://historico.tsj.gob.ve')}</u></font>", box_body_style)]
    ]
    
    box_table = Table(box_content, colWidths=[520])
    box_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#FAFAFA")),
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor("#CBD5E0")),
        ('INNERGRID', (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
        ('PADDING', (0,0), (-1,-1), 7),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(box_table)
    story.append(Spacer(1, 16))
    
    # 5. Full Decision Body Text (Formatted with TSJ Headings)
    texto = decision.get('texto_completo', '')
    paragraphs = texto.split('\n\n')
    
    for p in paragraphs:
        p_clean = p.strip()
        if not p_clean:
            continue
            
        # Detect if paragraph is a legal section heading (e.g. DE LA COMPETENCIA, ANTECEDENTES, DECISIÓN)
        if re.match(r'^(?:I+|II+|III+|IV+|V+|DE LA COMPETENCIA|MOTIVACIÓN|CONSIDERACIONES|ANTECEDENTES|DECISIÓN|DEL RECURSO INTERPUESTO)$', p_clean, re.IGNORECASE):
            story.append(Paragraph(f"<b>{p_clean.upper()}</b>", section_heading_style))
        else:
            # Escape HTML characters safely
            safe_text = p_clean.replace('<', '&lt;').replace('>', '&gt;')
            
            # Format key legal terms in bold like TSJ print
            safe_text = re.sub(r'\b(CON LUGAR|SIN LUGAR|ANULA|ORDENA|DECLARA|CONFIRMA|RECURSO DE CASACIÓN|RECURSO DE AMPARO|ACCESO INDEBIDO|DELITOS INFORMÁTICOS|PRUEBA DIGITAL|CADENA DE CUSTODIA)\b', r'<b>\1</b>', safe_text)
            
            story.append(Paragraph(safe_text, body_style))
            story.append(Spacer(1, 8))
            
    doc.build(story)


def export_summary_pdf(decisiones: List[Dict[str, Any]], filepath: str, title: str = "CATÁLOGO OFICIAL DE JURISPRUDENCIA - TSJ VENEZUELA") -> None:
    """Generates a summary PDF catalog containing all extracted decisions in TSJ print style."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=35, leftMargin=35, topMargin=35, bottomMargin=35)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Heading1'], fontName='Times-Bold',
        fontSize=14, leading=18, textColor=colors.HexColor("#000000"), alignment=TA_CENTER
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'], fontName='Times-Roman', fontSize=8.5, leading=11
    )
    cell_bold = ParagraphStyle(
        'CellBold', parent=styles['Normal'], fontName='Times-Bold', fontSize=8.5, leading=11, textColor=colors.HexColor("#000000")
    )
    
    story = []

    escudo_path = "assets/escudo_venezuela.png"
    if os.path.exists(escudo_path):
        img = Image(escudo_path, width=50, height=55)
        img.hAlign = 'CENTER'
        story.append(img)
        story.append(Spacer(1, 4))

    story.extend([
        Paragraph("REPÚBLICA BOLIVARIANA DE VENEZUELA - TRIBUNAL SUPREMO DE JUSTICIA", ParagraphStyle('H', fontName='Times-Bold', fontSize=10, alignment=TA_CENTER)),
        Paragraph(title, title_style),
        Paragraph(f"Total Registros Extraídos: {len(decisiones)} | Fecha de Emisión: {datetime.now().strftime('%d de %B de %Y')}", 
                  ParagraphStyle('Sub', fontName='Times-Italic', fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor("#4A5568"))),
        Spacer(1, 12)
    ])
    
    table_data = [
        [Paragraph("Sala TSJ", cell_bold), Paragraph("Nº Sent.", cell_bold), Paragraph("Expediente", cell_bold), Paragraph("Fecha", cell_bold), Paragraph("Ponente", cell_bold), Paragraph("Asunto / Síntesis", cell_bold)]
    ]
    
    for d in decisiones:
        table_data.append([
            Paragraph(str(d.get('sala', '')).replace("Sala de ", "").replace("Sala ", ""), cell_style),
            Paragraph(str(d.get('numero_sentencia', '')), cell_style),
            Paragraph(str(d.get('expediente', '')), cell_style),
            Paragraph(str(d.get('fecha', '')), cell_style),
            Paragraph(str(d.get('ponente', '')), cell_style),
            Paragraph(str(d.get('asunto', ''))[:75], cell_style),
        ])
        
    t = Table(table_data, colWidths=[80, 50, 75, 65, 95, 175])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#F7FAFC")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E0")),
        ('PADDING', (0,0), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(t)
    doc.build(story)